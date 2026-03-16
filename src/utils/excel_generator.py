"""
Generador de reportes en formato Excel con estilos profesionales.
"""
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


def _strip_tz(value):
    """Quita la zona horaria de un datetime/time para compatibilidad con openpyxl."""
    if hasattr(value, "tzinfo") and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _create_header_style(ws, titulo, subtitulo=""):
    """Agrega encabezado profesional al worksheet."""
    ws.merge_cells("A1:H1")
    header = ws["A1"]
    header.value = "UNIVERSIDAD NACIONAL DE TRUJILLO"
    header.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header.fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    header.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    ws.merge_cells("A2:H2")
    ws["A2"].value = titulo
    ws["A2"].font = Font(name="Calibri", size=14, bold=True, color="003D7A")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    if subtitulo:
        ws.merge_cells("A3:H3")
        ws["A3"].value = subtitulo
        ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="666666")
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 15
        return 4
    return 3


def _create_table_header(ws, row, columns):
    """Crea encabezado de tabla con estilos."""
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[row].height = 20


def _apply_table_style(ws, start_row, end_row, num_cols):
    """Aplica bordes y estilos a los datos de la tabla."""
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for row_idx in range(start_row, end_row + 1):
        is_alt = (row_idx - start_row) % 2 == 1
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if is_alt:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _add_footer(ws, row):
    """Agrega pie de página con fecha de generación."""
    ws.merge_cells(f"A{row}:H{row}")
    footer = ws[f"A{row}"]
    footer.value = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer.font = Font(name="Calibri", size=9, italic=True, color="999999")
    footer.alignment = Alignment(horizontal="right")


def _adjust_column_widths(ws, num_cols):
    """Ajusta automáticamente el ancho de columnas."""
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def generar_excel_pagos(pagos_data):
    """Genera un archivo Excel con reporte de pagos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"

    row = _create_header_style(ws, "REPORTE DE PAGOS Y VOUCHERS")
    row += 1

    columns = ["Código Voucher", "Usuario", "Concepto", "Monto (S/.)", "Fecha", "Pagado", "Validado"]
    _create_table_header(ws, row, columns)

    data_start_row = row + 1
    for pago in pagos_data:
        row += 1
        ws.cell(row=row, column=1).value = pago.get("codigo_voucher", "")
        ws.cell(row=row, column=2).value = pago.get("email", "")
        ws.cell(row=row, column=3).value = str(pago.get("concepto", "")).upper()
        ws.cell(row=row, column=4).value = float(pago.get("monto", 0))
        ws.cell(row=row, column=5).value = _strip_tz(pago.get("fecha_pago"))
        ws.cell(row=row, column=6).value = "Sí" if pago.get("pagado") else "No"
        ws.cell(row=row, column=7).value = "Sí" if pago.get("validado") else "No"

    _apply_table_style(ws, data_start_row, row, len(columns))
    _adjust_column_widths(ws, len(columns))

    ws.cell(row=row + 2, column=1).value = f"Total Ingresos: S/. {sum(float(p.get('monto', 0)) for p in pagos_data):,.2f}"
    ws.cell(row=row + 2, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row + 3, column=1).value = f"Total Registros: {len(pagos_data)}"
    ws.cell(row=row + 3, column=1).font = Font(bold=True, size=11)

    _add_footer(ws, row + 5)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_encuestas_resultados(encuestas_data):
    """Genera un archivo Excel con resultados de encuestas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    row = _create_header_style(ws, "REPORTE DE RESULTADOS DE ENCUESTAS")
    row += 1

    columns = ["Encuesta", "Pregunta", "Tipo Respuesta", "Respuesta", "Cantidad Respuestas", "Porcentaje"]
    _create_table_header(ws, row, columns)

    data_start_row = row + 1
    for item in encuestas_data:
        row += 1
        ws.cell(row=row, column=1).value = item.get("titulo_encuesta", "")
        ws.cell(row=row, column=2).value = item.get("texto_pregunta", "")
        ws.cell(row=row, column=3).value = item.get("tipo_respuesta", "")
        ws.cell(row=row, column=4).value = item.get("respuesta", "")
        ws.cell(row=row, column=5).value = int(item.get("cantidad", 0))
        ws.cell(row=row, column=6).value = f"{float(item.get('porcentaje', 0)):.1f}%"

    _apply_table_style(ws, data_start_row, row, len(columns))
    _adjust_column_widths(ws, len(columns))

    _add_footer(ws, row + 2)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_resultados_busqueda(resultados_data, titulo_reporte="REPORTE DE BÚSQUEDA"):
    """Genera un archivo Excel con resultados de búsqueda/consultas avanzadas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    row = _create_header_style(ws, titulo_reporte)
    row += 1

    if not resultados_data:
        ws.cell(row=row, column=1).value = "No hay resultados para mostrar"
        ws.cell(row=row, column=1).font = Font(italic=True)
        _add_footer(ws, row + 2)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    columns = list(resultados_data[0].keys()) if isinstance(resultados_data[0], dict) else []

    if not columns:
        columns = [f"Columna {i+1}" for i in range(len(resultados_data[0]))]

    _create_table_header(ws, row, columns)

    data_start_row = row + 1
    for item in resultados_data:
        row += 1
        for col_idx, col_name in enumerate(columns, 1):
            value = item.get(col_name) if isinstance(item, dict) else item[col_idx - 1]
            cell = ws.cell(row=row, column=col_idx)
            cell.value = _strip_tz(value)

    _apply_table_style(ws, data_start_row, row, len(columns))
    _adjust_column_widths(ws, len(columns))

    ws.cell(row=row + 2, column=1).value = f"Total Registros: {len(resultados_data)}"
    ws.cell(row=row + 2, column=1).font = Font(bold=True, size=11)

    _add_footer(ws, row + 4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
